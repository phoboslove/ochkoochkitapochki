"""Template renderer that respects canonical mappings + repeating-row tables.

Strategy:
  * Build a `context` dict from the BusinessContext + invoice instance,
    keyed by canonical placeholder names (company_name, total, items, …).
  * For DOCX: use docxtpl for placeholders + `{% for it in items %}` row blocks.
  * For XLSX: use openpyxl, do a global replace of `{{...}}` in cells and
    expand the *first* row that contains `{{items.*}}` into N rows.

The user maps real document labels to canonical keys; the renderer never
guesses semantics on its own.
"""
from __future__ import annotations

import copy
import re
from io import BytesIO
from typing import Any

from app.services.templates.placeholders import CANONICAL


_PLACEHOLDER_RE = re.compile(r"\{\{\s*([\w\.]+)\s*\}\}")


def render_template(
    *, content: bytes, format: str, context: dict[str, Any],
) -> tuple[bytes, str, str]:
    """Returns (bytes, mime, ext)."""
    if format == "docx":
        return _render_docx(content, context)
    if format == "xlsx":
        return _render_xlsx(content, context)
    # Fallback: treat as HTML text replacement.
    text = content.decode("utf-8", errors="ignore")
    out = _PLACEHOLDER_RE.sub(lambda m: str(_lookup(context, m.group(1))), text)
    return out.encode("utf-8"), "text/html", "html"


# ── DOCX ────────────────────────────────────────────────────────────────────

def _render_docx(content: bytes, context: dict[str, Any]) -> tuple[bytes, str, str]:
    from docxtpl import DocxTemplate
    tpl = DocxTemplate(BytesIO(content))
    tpl.render(context)
    out = BytesIO()
    tpl.save(out)
    return (out.getvalue(),
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "docx")


# ── XLSX ────────────────────────────────────────────────────────────────────

def _render_xlsx(content: bytes, context: dict[str, Any]) -> tuple[bytes, str, str]:
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(content))

    items = context.get("items") or []
    for ws in wb.worksheets:
        # 1) Find a row referencing {{items.*}} → expand into len(items) rows.
        row_template_idx = _find_items_row(ws)
        if row_template_idx is not None and items:
            _expand_items_rows(ws, row_template_idx, items, context)

        # 2) Replace remaining scalar placeholders in all string cells.
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{{" in cell.value:
                    cell.value = _PLACEHOLDER_RE.sub(
                        lambda m: str(_lookup(context, m.group(1))),
                        cell.value,
                    )

    out = BytesIO()
    wb.save(out)
    return (out.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx")


def _find_items_row(ws) -> int | None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "{{items." in cell.value:
                return cell.row
    return None


def _expand_items_rows(ws, row_idx: int, items: list[dict], context: dict) -> None:
    """Duplicate the template row into N rows, substituting items.*. Preserves
    formatting via openpyxl row attribute copy."""
    template_cells = [(c.column_letter, c.value, _copy_style(c)) for c in ws[row_idx]]
    needed = len(items)
    if needed == 0:
        return
    # Insert needed-1 blank rows below the template row.
    if needed > 1:
        ws.insert_rows(row_idx + 1, amount=needed - 1)

    for offset, it in enumerate(items):
        r = row_idx + offset
        ctx_row = {**context, "it": it, "idx": offset + 1}
        for col_letter, value, style in template_cells:
            cell = ws[f"{col_letter}{r}"]
            if isinstance(value, str) and "{{" in value:
                rendered = _PLACEHOLDER_RE.sub(lambda m: str(_lookup(ctx_row, m.group(1))), value)
                cell.value = rendered
            else:
                cell.value = value
            _apply_style(cell, style)


def _copy_style(cell) -> dict:
    return {
        "font": copy.copy(cell.font),
        "fill": copy.copy(cell.fill),
        "border": copy.copy(cell.border),
        "alignment": copy.copy(cell.alignment),
        "number_format": cell.number_format,
    }


def _apply_style(cell, style: dict) -> None:
    cell.font = style["font"]
    cell.fill = style["fill"]
    cell.border = style["border"]
    cell.alignment = style["alignment"]
    cell.number_format = style["number_format"]


# ── Lookup ──────────────────────────────────────────────────────────────────

def _lookup(context: dict[str, Any], dotted: str) -> Any:
    cur: Any = context
    for part in dotted.split("."):
        if cur is None: return ""
        cur = cur.get(part) if isinstance(cur, dict) else getattr(cur, part, "")
    return "" if cur is None else cur


# ── Demo context builder (for preview render) ───────────────────────────────

def build_demo_context() -> dict[str, Any]:
    """Materialize the canonical sample values as a single nested context.
    The canonical keys are flat — they are duplicated into nested form too so
    DOCX templates can use both `{{company_name}}` and `{{company.name}}`."""
    flat = {k: f.sample for k, f in CANONICAL.items()}
    flat["items"] = [
        {"name": "Consulting services", "qty": 8, "price": 30000, "total": 240000},
        {"name": "Implementation",       "qty": 1, "price": 210000, "total": 210000},
    ]
    return {
        **flat,
        "company": {"name": flat["company_name"], "bin": flat["company_bin"],
                    "address": flat["company_address"]},
        "client":  {"name": flat["client_name"], "bin": flat["client_bin"],
                    "address": flat["client_address"]},
        "invoice": {"number": flat["invoice_number"], "date": flat["invoice_date"],
                    "currency": flat["currency"], "subtotal": flat["subtotal"],
                    "vat": flat["vat"], "total": flat["total"]},
    }


def build_context_from_template(tpl_mappings: dict[str, dict], context: dict[str, Any]) -> dict[str, Any]:
    """Apply the user-confirmed mappings: re-key the context to whatever
    source labels the template uses (e.g. `{{Поставщик}}` → company_name).

    For now both the canonical and the source-label keys are exposed so the
    template can use either form. Mostly relevant for DOCX with literal RU/KZ
    placeholder names — the user may have created them with foreign aliases.
    """
    out = {**context}
    for canonical, m in (tpl_mappings or {}).items():
        if not m.get("confirmed"): continue
        label = m.get("source_label")
        if label and canonical in context:
            # Sanitize the label to a valid Jinja token (e.g. "Поставщик" stays as-is for docxtpl).
            out[label.strip()] = context[canonical]
    return out
